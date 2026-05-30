import csv
import re
from io import BytesIO
from django.http import HttpResponse
from django.db import transaction
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import unicodedata
import logging
import xlrd
import chardet

logger = logging.getLogger(__name__)


def normalize_header(value):
    """Normalize header for flexible matching"""
    if value is None:
        return ''
    header = str(value).strip().lower()
    # Remove accents
    header = ''.join(c for c in unicodedata.normalize('NFD', header)
                     if unicodedata.category(c) != 'Mn')
    # Remove special characters and normalize spaces
    header = re.sub(r'[\s_\-]+', '', header)
    header = re.sub(r'[^a-z0-9]', '', header)
    return header


class ColumnMapper:
    """Intelligent column mapping system for model fields"""
    
    # Field aliases for flexible matching - EXTENSIVE coverage
    FIELD_ALIASES = {
        # Student fields - EXTENSIVE aliases for better detection
        'cin': ['cin', 'cni', 'cardeid', 'carte identité', 'cartenumero', 'id number', 'numid', 'numero id', 
                'numéro id', 'num id', 'numero cin', 'id cin', 'carte numero', 'num carte', 'identifiant',
                'num identifiant', 'code id', 'id code', 'numero identite', 'num identite'],
        
        'nom_fr': ['nom', 'nomfr', 'lastname', 'nom fr', 'name', 'nomfamilial', 'famille', 'nom famille',
                   'nom etudiant', 'nom student', 'last name', 'nom personnel', 'surname', 'nom de famille',
                   'family name', 'nom complet', 'full name', 'prenom nom'],
        
        'prenom_fr': ['prenom', 'premnom', 'firstname', 'prenom fr', 'prenom nom', 'given name', 'prenom etudiant',
                      'first name', 'nom personnel', 'prenom principal', 'forename', 'first', 'given'],
        
        'email': ['email', 'mail', 'e-mail', 'adressemail', 'adresse email', 'email address', 'adresse mail',
                  'electronic mail', 'email etudiant', 'student email', 'courrier', 'adresse electronique'],
        
        'numTel': ['tel', 'telephone', 'phone', 'gsm', 'mobile', 'numtel', 'numero', 'numéro', 'tel phone',
                   'numero telephone', 'num telephone', 'numero gsm', 'num gsm', 'portable', 'phone number',
                   'numero mobile', 'num mobile', 'telephone number', 'contact', 'numero contact'],
        
        'dateNaissance': ['datenaissance', 'datedenaissance', 'naissance', 'birthdate', 'date birth', 
                          'date naissance', 'date de naissance', 'date birth', 'birth date', 'date of birth',
                          'date nascimento', 'date naissance etudiant', 'dob', 'birth', 'naissance date'],
        
        'adresse': ['adresse', 'address', 'rue', 'lieu', 'adresse lieu', 'full address', 'street address',
                    'home address', 'residential address', 'adresse residence', 'adresse domicile', 'location'],
        
        'nationalite': ['nationalite', 'nationality', 'pays', 'country', 'nation', 'nationalité', 'pays origine'],
        
        'passport': ['passport', 'passeport', 'passport number', 'num passport', 'numero passeport', 'passport no'],
        
        'genre': ['genre', 'sex', 'gender', 'sexe', 'h/f', 'male/female', 'genre etudiant', 'gender etudiant'],
        
        'groupe': ['groupe', 'group', 'class', 'classe', 'section', 'group number', 'numero groupe', 'classe etudiant'],
        
        'dateInscription': ['dateinscription', 'inscription', 'date inscription', 'enroll date', 'enrollment date'],
        
        'situation_s5': ['situation s5', 'situations5', 'situa5', 'situation'],
        
        'situation_pfe': ['situation pfe', 'situationpfe', 'pfe situation'],
        
        'licence': ['licence', 'degree', 'mention', 'parcours', 'diplome', 'license', 'licence name'],
        
        'specialite': ['specialite', 'speciality', 'spec', 'specialisation', 'specialty', 'specialization'],
        
        # Teacher fields
        'matricule': ['matricule', 'matricul', 'id enseignant', 'idenseignant', 'teacher id', 'employee id',
                      'numero matricule', 'num matricule', 'id teacher'],
        
        'nom': ['nom', 'lastname', 'name', 'nomfamilial', 'famille', 'nom famille', 'last name'],
        
        'prenom': ['prenom', 'prénom', 'firstname', 'given name', 'first name'],
        
        'numtel': ['tel', 'telephone', 'téléphone', 'phone', 'gsm', 'mobile', 'numtel', 'numero', 'numéro'],
        
        'grade': ['grade', 'rank', 'titre', 'fonction', 'position', 'rang'],
        
        'daterecrutement': ['daterecrutement', 'date recrutement', 'date recruitment', 'recruitment date', 
                           'date embauche', 'date hiring', 'hire date'],
        
        'statutadministratif': ['statutadministratif', 'statut administratif', 'administrative status', 'status'],
        
        'departement': ['departement', 'department', 'dept', 'departement name'],
        
        'role': ['role', 'fonction', 'position', 'type role', 'role type'],
    }
    
    def __init__(self, excel_headers_dict):
        """
        Initialize mapper with Excel headers
        excel_headers_dict: {normalized_header: column_index}
        """
        self.excel_headers = excel_headers_dict
        self.mapping = {}  # normalized_header -> (model_field, original_header)
        self.mapped_fields = set()  # model fields that were mapped
        self.ignored_headers = set()  # headers that couldn't be mapped
        self.auto_map()
    
    def auto_map(self):
        """Automatically map Excel headers to model fields"""
        for excel_header in self.excel_headers.keys():
            matched_field = self._find_best_match(excel_header)
            if matched_field:
                self.mapping[excel_header] = matched_field
                self.mapped_fields.add(matched_field)
            else:
                self.ignored_headers.add(excel_header)
    
    def _find_best_match(self, normalized_header):
        """Find best matching model field for an Excel header"""
        # Scoring-based match: exact match > startswith/endswith > substring
        best_score = 0
        best_field = None

        for model_field, aliases in self.FIELD_ALIASES.items():
            for alias in aliases:
                normalized_alias = normalize_header(alias)
                if not normalized_alias:
                    continue

                score = 0
                if normalized_header == normalized_alias:
                    score = 100
                elif normalized_header.startswith(normalized_alias) or normalized_header.endswith(normalized_alias) or normalized_alias.startswith(normalized_header) or normalized_alias.endswith(normalized_header):
                    score = 50
                elif normalized_alias in normalized_header or normalized_header in normalized_alias:
                    score = 10

                if score > best_score:
                    best_score = score
                    best_field = model_field

                # immediate return on exact match
                if score == 100:
                    return model_field

        return best_field
    
    def get_mapped_columns(self):
        """Return mapping of model fields to Excel headers"""
        result = {}
        for excel_header, model_field in self.mapping.items():
            result[model_field] = excel_header
        return result
    
    def get_ignored_columns(self):
        """Return list of ignored Excel headers"""
        return list(self.ignored_headers)
    
    def get_missing_fields(self):
        """Return list of model fields that weren't found in Excel"""
        return [field for field in self.FIELD_ALIASES.keys() if field not in self.mapped_fields]
    
    def map_row(self, row_data):
        """Map a row from Excel column names to model field names"""
        mapped_row = {}
        for excel_header, value in row_data.items():
            if excel_header in self.mapping:
                model_field = self.mapping[excel_header]
                mapped_row[model_field] = value
        return mapped_row


class DataCleaner:
    """Clean and normalize imported data"""
    
    @staticmethod
    def normalize_spaces(value):
        """Normalize whitespace"""
        if value is None or value == '':
            return ''
        return str(value).strip().replace('\r\n', ' ').replace('\n', ' ')
    
    @staticmethod
    def remove_accents(value):
        """Remove accents from string"""
        if not isinstance(value, str):
            return value
        return ''.join(c for c in unicodedata.normalize('NFD', value)
                      if unicodedata.category(c) != 'Mn')
    
    @staticmethod
    def clean_cin(value):
        """Clean CIN - keep only digits"""
        if not value:
            return None
        value = str(value).strip()
        # Remove any non-digit characters
        cleaned = re.sub(r'\D', '', value)
        return cleaned if cleaned else None
    
    @staticmethod
    def clean_phone(value):
        """Clean phone number - keep only digits"""
        if not value:
            return None
        value = str(value).strip()
        # Try to extract digits
        digits = re.sub(r'\D', '', value)
        # Remove leading zero if present (to normalize)
        if digits.startswith('0'):
            digits = digits[1:]
        return digits if digits else None
    
    @staticmethod
    def clean_email(value):
        """Clean email"""
        if not value:
            return None
        return str(value).strip().lower()
    
    @staticmethod
    def parse_date(value):
        """Parse date from various formats including datetime strings"""
        if not value:
            return None
        
        value_str = str(value).strip()
        
        # Handle datetime objects that have been converted to string (e.g., "1995-05-15 00:00:00")
        if ' ' in value_str and ':' in value_str:
            # This looks like a datetime string, extract just the date part
            date_part = value_str.split(' ')[0]
            # Try to parse the date part
            try:
                datetime.strptime(date_part, '%Y-%m-%d')
                return date_part
            except:
                value_str = date_part  # Continue with the date part
        
        # Handle Excel date numbers (serial dates)
        try:
            if isinstance(value, (int, float)) and 0 < value < 100000:
                # This is likely an Excel serial date
                excel_epoch = datetime(1899, 12, 30)
                delta = datetime.timedelta(days=value)
                date_obj = excel_epoch + delta
                return date_obj.strftime('%Y-%m-%d')
        except:
            pass
        
        # Try common date formats
        date_formats = [
            '%Y-%m-%d',      # 2025-05-22
            '%d-%m-%Y',      # 22-05-2025
            '%d/%m/%Y',      # 22/05/2025
            '%d/%m/%y',      # 22/05/25
            '%Y/%m/%d',      # 2025/05/22
            '%d.%m.%Y',      # 22.05.2025
            '%d-%m-%y',      # 22-05-25
        ]
        
        for fmt in date_formats:
            try:
                date_obj = datetime.strptime(value_str, fmt)
                return date_obj.strftime('%Y-%m-%d')
            except ValueError:
                continue
        
        return value_str  # Return as-is if can't parse
    
    @staticmethod
    def clean_value(value, field_name):
        """Clean a value based on field type"""
        if value is None or value == '':
            return None
        
        field_name_lower = field_name.lower()
        
        if field_name_lower == 'cin':
            return DataCleaner.clean_cin(value)
        elif field_name_lower == 'numtel':
            return DataCleaner.clean_phone(value)
        elif field_name_lower == 'email':
            return DataCleaner.clean_email(value)
        elif field_name_lower in ['datenaissance', 'daterecrutement']:
            return DataCleaner.parse_date(value)
        elif field_name_lower in ['nom_fr', 'prenom_fr', 'adresse', 'nationalite']:
            return DataCleaner.normalize_spaces(value)
        elif field_name_lower == 'genre':
            val = str(value).strip().upper()
            return 'F' if val in ['F', 'FEMALE', 'FEMME'] else 'M'
        else:
            return DataCleaner.normalize_spaces(value)
    
    @staticmethod
    def clean_row(row_data, field_names):
        """Clean an entire row"""
        cleaned = {}
        for field_name, value in row_data.items():
            cleaned[field_name] = DataCleaner.clean_value(value, field_name)
        return cleaned


class ImportReporter:
    """Generate detailed import reports"""
    
    def __init__(self):
        self.imported_rows = []
        self.errors = {}  # row_idx -> list of errors
        self.warnings = {}  # row_idx -> list of warnings
        self.skipped_rows = []
    
    def add_imported_row(self, row_idx, row_data):
        """Record a successfully imported row"""
        self.imported_rows.append({
            'row_idx': row_idx,
            'data': row_data
        })
    
    def add_error(self, row_idx, error_msg):
        """Add error for a row"""
        if row_idx not in self.errors:
            self.errors[row_idx] = []
        self.errors[row_idx].append(error_msg)
    
    def add_warning(self, row_idx, warning_msg):
        """Add warning for a row"""
        if row_idx not in self.warnings:
            self.warnings[row_idx] = []
        self.warnings[row_idx].append(warning_msg)
    
    def add_skipped_row(self, row_idx, reason):
        """Record a skipped row"""
        self.skipped_rows.append({
            'row_idx': row_idx,
            'reason': reason
        })
    
    def get_report(self):
        """Get detailed import report"""
        return {
            'total_processed': len(self.imported_rows) + len(self.errors) + len(self.skipped_rows),
            'imported_count': len(self.imported_rows),
            'error_count': len(self.errors),
            'skipped_count': len(self.skipped_rows),
            'errors': self.errors,
            'warnings': self.warnings,
            'skipped_rows': self.skipped_rows,
            'success': len(self.errors) == 0
        }


def detect_file_type_and_content(file_obj):
    """
    Detect actual file type by reading bytes and extension
    Returns: (detected_type, file_bytes, error_message)
    detected_type: 'xlsx', 'xls', 'csv', or 'unknown'
    """
    try:
        file_obj.seek(0)
        file_bytes = file_obj.read()
        
        if not file_bytes:
            return 'unknown', file_bytes, "Fichier vide"
        
        filename = getattr(file_obj, 'name', '').lower()
        extension = filename.split('.')[-1] if '.' in filename else ''
        
        # Check by file signature (magic bytes)
        logger.info(f"📋 Détection du fichier: {filename}")
        logger.info(f"🔍 Extension: .{extension}")
        
        # XLSX/DOCX detection: PK signature (ZIP format)
        if file_bytes[:2] == b'PK':
            logger.info("✓ Signature ZIP détectée (XLSX/Office)")
            # XLSX has specific structure, check more carefully
            if extension in ['xlsx', 'xls'] or extension == '':
                return 'xlsx', file_bytes, None
            return 'xlsx', file_bytes, None
        
        # XLS detection: BIFF signature
        if file_bytes[:8] in [b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1', b'\xfd\x37\x37\x37\x37\x37\x37\x37']:
            logger.info("✓ Signature XLS (BIFF5/BIFF8) détectée")
            return 'xls', file_bytes, None
        
        # CSV detection: try to decode and check structure
        # First, detect encoding
        detection = chardet.detect(file_bytes)
        encoding = detection.get('encoding', 'utf-8') or 'utf-8'
        
        logger.info(f"🔤 Encodage détecté: {encoding}")
        
        try:
            content = file_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            # Fallback encodings
            for fallback_enc in ['utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    content = file_bytes.decode(fallback_enc)
                    encoding = fallback_enc
                    logger.info(f"✓ Décodage réussi avec {fallback_enc}")
                    break
                except:
                    continue
            else:
                return 'unknown', file_bytes, "Impossible de décoder le fichier (encodage incompatible)"
        
        # Check if content looks like CSV or plain text
        lines = content.split('\n')[:5]  # Check first 5 lines
        if lines and len(lines[0]) > 0:
            # It's probably text-based (CSV or TXT)
            logger.info("✓ Format texte détecté (CSV ou TXT)")
            return 'csv', file_bytes, None
        
        return 'unknown', file_bytes, "Format de fichier non reconnu"
        
    except Exception as e:
        logger.error(f"❌ Erreur détection fichier: {str(e)}")
        return 'unknown', b'', f"Erreur lors de la détection: {str(e)}"


def read_xlsx_file(file_bytes):
    """
    Read XLSX file (.xlsx)
    Returns: (headers_dict, rows_list, errors_list, column_mapper)
    """
    try:
        logger.info("📖 Lecture XLSX avec openpyxl...")
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        
        if not ws or ws.max_row < 2:
            logger.warning("⚠️  Fichier XLSX vide")
            return {}, [], ['Fichier Excel vide'], None
        
        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value is not None:
                header = normalize_header(str(cell.value))
                if header:  # Only add non-empty headers
                    headers[header] = col_idx
        
        if not headers:
            logger.warning("⚠️  Aucun en-tête trouvé")
            return {}, [], ['Aucun en-tête trouvé dans le fichier'], None
        
        logger.info(f"✓ {len(headers)} colonnes détectées")
        
        # Create column mapper
        column_mapper = ColumnMapper(headers)
        
        # Get data rows
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for cell in ws[row_idx]:
                if cell.column in headers.values():
                    # Find the header name for this column
                    for header_name, col_idx in headers.items():
                        if col_idx == cell.column:
                            if cell.value is not None:
                                row_data[header_name] = cell.value
                            break
            
            # Skip completely empty rows
            if any(row_data.values()):
                rows.append((row_idx, row_data))
        
        logger.info(f"✓ {len(rows)} lignes de données trouvées")
        return headers, rows, [], column_mapper
        
    except Exception as e:
        error_msg = f"Erreur lecture XLSX: {str(e)}"
        logger.error(f"❌ {error_msg}")
        return {}, [], [error_msg], None


def read_xls_file(file_bytes):
    """
    Read XLS file (.xls) using xlrd
    Returns: (headers_dict, rows_list, errors_list, column_mapper)
    """
    try:
        logger.info("📖 Lecture XLS avec xlrd...")
        book = xlrd.open_workbook(file_contents=file_bytes, on_demand=True)
        sheet = book.sheet_by_index(0)
        
        if sheet.nrows < 2:
            logger.warning("⚠️  Fichier XLS vide")
            return {}, [], ['Fichier XLS vide'], None
        
        # Get headers from first row
        headers = {}
        for col_idx in range(sheet.ncols):
            header_value = sheet.cell_value(0, col_idx)
            if header_value:
                header = normalize_header(str(header_value))
                if header:  # Only add non-empty headers
                    headers[header] = col_idx
        
        if not headers:
            logger.warning("⚠️  Aucun en-tête trouvé")
            return {}, [], ['Aucun en-tête trouvé'], None
        
        logger.info(f"✓ {len(headers)} colonnes détectées")
        
        # Create column mapper
        column_mapper = ColumnMapper(headers)
        
        # Build a reverse mapping from column index -> normalized header name
        colidx_to_header = {col_idx: header_name for header_name, col_idx in headers.items()}

        # Helper to decide if a header likely represents a date
        def _is_date_header(hname):
            if not hname:
                return False
            return any(k in hname for k in ('date', 'datenaissance', 'datedenaissance', 'daterecrutement', 'dateinscription', 'dob'))

        # Get data rows
        rows = []
        for row_idx in range(1, sheet.nrows):
            row_data = {}
            for col_idx in headers.values():
                value = sheet.cell_value(row_idx, col_idx)
                header_name = colidx_to_header.get(col_idx)

                # Only attempt Excel-date conversion when the header looks like a date field
                if header_name and isinstance(value, float) and _is_date_header(header_name):
                    try:
                        from xlrd import xldate_as_datetime
                        value = xldate_as_datetime(value, book.datemode)
                    except:
                        pass

                # Only set the value if it's truthy (non-empty) to avoid adding empty keys
                if value or (isinstance(value, (int, float)) and value == 0):
                    row_data[header_name] = value

            # Skip completely empty rows
            if any(v for v in row_data.values() if v is not None and str(v).strip() != ''):
                # The API expects 1-based Excel row numbers
                rows.append((row_idx + 1, row_data))
        
        logger.info(f"✓ {len(rows)} lignes de données trouvées")
        return headers, rows, [], column_mapper
        
    except Exception as e:
        error_msg = f"Erreur lecture XLS: {str(e)}"
        logger.error(f"❌ {error_msg}")
        return {}, [], [error_msg], None


def detect_csv_separator(content_sample):
    """
    Detect CSV separator by analyzing first line
    Returns: (separator, encoding)
    """
    separators = [',', ';', '\t', '|']
    
    lines = content_sample.split('\n')[:2]
    if not lines:
        return ',', 'utf-8'
    
    first_line = lines[0]
    
    # Count occurrences of each separator
    sep_counts = {sep: first_line.count(sep) for sep in separators}
    
    # Choose the one with most occurrences (but at least 1)
    valid_seps = {sep: count for sep, count in sep_counts.items() if count > 0}
    
    if valid_seps:
        detected_sep = max(valid_seps.items(), key=lambda x: x[1])[0]
        logger.info(f"✓ Séparateur CSV détecté: '{detected_sep}' ({valid_seps[detected_sep]} occurrences)")
        return detected_sep, 'utf-8'
    
    logger.warning("⚠️  Aucun séparateur trouvé, utilisation du séparateur par défaut")
    return ',', 'utf-8'


def read_csv_file_robust(file_bytes):
    """
    Read CSV file with robust separator and encoding detection
    Returns: (headers_dict, rows_list, errors_list, column_mapper)
    """
    try:
        logger.info("📖 Lecture CSV...")
        
        # Detect encoding
        detection = chardet.detect(file_bytes)
        encoding = detection.get('encoding', 'utf-8') or 'utf-8'
        logger.info(f"🔤 Encodage détecté: {encoding}")
        
        # Try primary encoding, then fallbacks
        content = None
        for enc in [encoding, 'utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                content = file_bytes.decode(enc)
                if encoding != enc:
                    logger.info(f"✓ Décodage réussi avec {enc}")
                break
            except:
                continue
        
        if content is None:
            return {}, [], ['Impossible de décoder le fichier CSV'], None
        
        # Detect separator
        separator, detected_enc = detect_csv_separator(content)
        
        # Parse CSV
        reader = csv.DictReader(content.splitlines(), delimiter=separator)
        
        if not reader.fieldnames:
            logger.warning("⚠️  Fichier CSV vide")
            return {}, [], ['Fichier CSV vide'], None
        
        # Normalize headers
        headers = {}
        for header in reader.fieldnames:
            if header:  # Skip None or empty headers
                normalized = normalize_header(header)
                if normalized:
                    headers[normalized] = len(headers) + 1
        
        if not headers:
            logger.warning("⚠️  Aucun en-tête valide trouvé")
            return {}, [], ['Aucun en-tête trouvé'], None
        
        logger.info(f"✓ {len(headers)} colonnes détectées")
        
        # Create column mapper
        column_mapper = ColumnMapper(headers)
        
        # Get data rows
        rows = []
        for row_idx, row in enumerate(reader, start=2):
            row_data = {}
            for original_header in reader.fieldnames:
                if original_header:
                    normalized_header = normalize_header(original_header)
                    value = row.get(original_header)
                    if value is not None:
                        value_str = str(value).strip()
                        if value_str:
                            row_data[normalized_header] = value_str
            
            # Skip completely empty rows
            if any(row_data.values()):
                rows.append((row_idx, row_data))
        
        logger.info(f"✓ {len(rows)} lignes de données trouvées")
        return headers, rows, [], column_mapper
        
    except Exception as e:
        error_msg = f"Erreur lecture CSV: {str(e)}"
        logger.error(f"❌ {error_msg}")
        return {}, [], [error_msg], None


def read_file_intelligent(file_obj):
    """
    Universal file reader - detects format and reads accordingly
    Supports: .xlsx, .xls, .csv
    Returns: (headers_dict, rows_list, errors_list, column_mapper)
    """
    try:
        logger.info("🚀 Début détection et lecture de fichier...")
        
        # Step 1: Detect file type
        detected_type, file_bytes, detect_error = detect_file_type_and_content(file_obj)
        
        if detect_error:
            logger.error(f"❌ Erreur détection: {detect_error}")
            return {}, [], [detect_error], None
        
        logger.info(f"📋 Type détecté: {detected_type.upper()}")
        
        # Step 2: Read file with appropriate reader
        if detected_type == 'xlsx':
            logger.info("→ Utilisation du lecteur XLSX...")
            headers, rows, errors, mapper = read_xlsx_file(file_bytes)
            if errors and 'openpyxl' not in str(errors[0]).lower():
                # openpyxl failed, try CSV fallback
                logger.warning("⚠️  Lecture XLSX échouée, tentative CSV...")
                headers, rows, errors, mapper = read_csv_file_robust(file_bytes)
            return headers, rows, errors, mapper
        
        elif detected_type == 'xls':
            logger.info("→ Utilisation du lecteur XLS...")
            return read_xls_file(file_bytes)
        
        elif detected_type == 'csv':
            logger.info("→ Utilisation du lecteur CSV...")
            return read_csv_file_robust(file_bytes)
        
        else:
            # Fallback: try all readers in order
            logger.warning("⚠️  Type inconnu, tentative de lecture multi-format...")
            
            # Try XLSX first
            headers, rows, errors, mapper = read_xlsx_file(file_bytes)
            if rows:
                return headers, rows, errors, mapper
            
            # Try XLS
            headers, rows, errors, mapper = read_xls_file(file_bytes)
            if rows:
                return headers, rows, errors, mapper
            
            # Try CSV
            headers, rows, errors, mapper = read_csv_file_robust(file_bytes)
            if rows:
                return headers, rows, errors, mapper
            
            return {}, [], ['Impossible de lire le fichier - format non supporté'], None
    
    except Exception as e:
        error_msg = f"Erreur fatale lecture fichier: {str(e)}"
        logger.error(f"❌ {error_msg}")
        return {}, [], [error_msg], None


def read_excel_file(file_obj):
    """
    Legacy function - delegates to intelligent reader
    Maintained for backward compatibility
    """
    return read_file_intelligent(file_obj)


def read_csv_file(file_obj):
    """
    Legacy function - delegates to intelligent reader
    Maintained for backward compatibility
    """
    return read_file_intelligent(file_obj)




def validate_required_fields(row_data, required_fields):
    """
    Validate that required fields are present and not empty
    Returns: (is_valid, missing_fields)
    """
    missing = []
    for field in required_fields:
        value = row_data.get(field)
        if value is None or (isinstance(value, str) and not value.strip()):
            missing.append(field)
    
    return len(missing) == 0, missing


def validate_required_fields_flexible(row_data, required_fields, mapped_fields):
    """
    Validate required fields with flexible checking.
    Only validates fields that were actually found in the Excel file.
    
    Args:
        row_data: cleaned row data
        required_fields: list of field names that are absolutely required
        mapped_fields: set of fields that were found in the Excel file
    
    Returns:
        (is_valid, missing_fields_in_data)
    """
    missing_in_data = []
    
    for field in required_fields:
        # Only validate if field was in the Excel file
        if field not in mapped_fields:
            continue
        
        value = row_data.get(field)
        if value is None or (isinstance(value, str) and not value.strip()):
            missing_in_data.append(field)
    
    return len(missing_in_data) == 0, missing_in_data
